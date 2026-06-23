from flask import Flask, render_template_string, request, jsonify, session, send_file, redirect, url_for
import csv
from collections import defaultdict
import os
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Change for production

PRODUCTS_FILE = 'products.json'

DEFAULT_PRODUCTS = {
    '2 in 1 Stylus Pen': {'cost': 100, 'keywords': ['Stylus']},
    'Stylus Pen': {'cost': 650, 'keywords': ['Stylus Pen']},
    '2-in-1 Card Reader': {'cost': 170, 'keywords': ['2 In1']},
    '3-in-1 Card Reader': {'cost': 170, 'keywords': ['Memory Card Reader']},
    'Circuit Breaker': {'cost': 1800, 'keywords': ['Circuit Breaker']},
    'Smart Plug': {'cost': 1100, 'keywords': ['Smart Socket']},
    'Otg Adapter': {'cost': 170, 'keywords': ['Otg Adapter', 'OTG']},
}

STORE_MAP = {
    'PK2NBYJ4S19': 'Global Gadgets',
    'PK2NBXPB2G6': 'Na Imports',
    'PK2NBODNVCR': 'Quality Club',
}

def load_products():
    if os.path.exists(PRODUCTS_FILE):
        with open(PRODUCTS_FILE, 'r') as f:
            return json.load(f)
    else:
        save_products(DEFAULT_PRODUCTS)
        return DEFAULT_PRODUCTS

def save_products(products):
    with open(PRODUCTS_FILE, 'w') as f:
        json.dump(products, f, indent=4)

# ---------- Helper functions ----------
def categorize(product_name, price_paid, net, products):
    if net < 0:
        return 'Other'
    if 'Stylus' in product_name or 'Stylus Pen' in product_name:
        return '2 in 1 Stylus Pen' if price_paid < 700 else 'Stylus Pen'
    for cat, data in products.items():
        for keyword in data.get('keywords', []):
            if keyword.lower() in product_name.lower():
                return cat
    return 'Other'

def get_cost(category, products):
    return products.get(category, {}).get('cost', 0)

def process_file(filepath):
    products = load_products()
    orders = {}
    period = "Unknown"
    store = "Unknown"
    try:
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter=';')
            for row in reader:
                if period == "Unknown" and 'Statement Period' in row:
                    period = row['Statement Period']
                if store == "Unknown" and 'Short Code' in row:
                    raw_store = row['Short Code']
                    store = STORE_MAP.get(raw_store, raw_store)
                order_id = row['Order Line ID']
                amount = float(row['Amount(Include Tax)'].replace(',', '').strip())
                fee_name = row['Fee Name']
                product_name = row['Product Name']
                if order_id not in orders:
                    orders[order_id] = {'total': 0.0, 'product_name': product_name, 'price_paid': 0.0}
                orders[order_id]['total'] += amount
                if fee_name == 'Product Price Paid by Buyer':
                    orders[order_id]['price_paid'] = amount
                if product_name:
                    orders[order_id]['product_name'] = product_name
    except Exception as e:
        return None, str(e)

    positive_payouts = defaultdict(list)
    negative_payouts = defaultdict(list)
    all_categories = set()

    for data in orders.values():
        net = data['total']
        product = data['product_name']
        price = data['price_paid']
        cat = categorize(product, price, net, products)
        all_categories.add(cat)
        if net < 0:
            negative_payouts[cat].append(net)
        else:
            positive_payouts[cat].append(net)

    sum_pos = sum(sum(lst) for lst in positive_payouts.values())
    sum_neg = sum(sum(lst) for lst in negative_payouts.values())
    total_payout = sum_pos + sum_neg

    total_positive_profit = 0.0
    category_positive_profit = {}
    for cat in all_categories:
        pos_list = positive_payouts.get(cat, [])
        count_pos = len(pos_list)
        sum_pos_cat = sum(pos_list)
        cost = get_cost(cat, products)
        gross_profit = sum_pos_cat - (count_pos * cost)
        category_positive_profit[cat] = gross_profit
        total_positive_profit += gross_profit

    net_profit = total_positive_profit + sum_neg

    return {
        'period': period,
        'store': store,
        'positive_payouts': positive_payouts,
        'negative_payouts': negative_payouts,
        'category_positive_profit': category_positive_profit,
        'total_positive_profit': total_positive_profit,
        'total_negative': sum_neg,
        'total_payout': total_payout,
        'net_profit': net_profit,
    }, None

# ---------- Export functions ----------
def generate_excel_summary(results, summary, grand_total, loss, final_profit):
    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("Per-File Results")
    row = 1
    for r in results:
        ws1.cell(row=row, column=1, value="File: " + r['filename'])
        ws1.cell(row=row, column=2, value="Period: " + r['period'])
        ws1.cell(row=row, column=3, value="Store: " + r['store'])
        row += 1
        ws1.cell(row=row, column=1, value="Positive Payouts")
        row += 1
        for cat, data in r['positive'].items():
            ws1.cell(row=row, column=1, value=cat)
            ws1.cell(row=row, column=2, value="orders: " + str(data['count']))
            ws1.cell(row=row, column=3, value=round(data['total'], 2))
            row += 1
        ws1.cell(row=row, column=1, value="Negative Payouts")
        row += 1
        for cat, data in r['negative'].items():
            ws1.cell(row=row, column=1, value=cat)
            ws1.cell(row=row, column=2, value="orders: " + str(data['count']))
            ws1.cell(row=row, column=3, value=round(data['total'], 2))
            row += 1
        ws1.cell(row=row, column=1, value="Gross Profit (after cost)")
        row += 1
        for cat, profit in r['profits'].items():
            ws1.cell(row=row, column=1, value=cat)
            ws1.cell(row=row, column=2, value=round(profit, 2))
            row += 1
        ws1.cell(row=row, column=1, value="Total Gross Profit")
        ws1.cell(row=row, column=2, value=round(r['total_gross'], 2))
        row += 1
        ws1.cell(row=row, column=1, value="Total Negative Payouts")
        ws1.cell(row=row, column=2, value=round(r['total_neg'], 2))
        row += 1
        ws1.cell(row=row, column=1, value="Total Payout (before cost)")
        ws1.cell(row=row, column=2, value=round(r['total_payout'], 2))
        row += 1
        ws1.cell(row=row, column=1, value="Net Profit (after cost)")
        ws1.cell(row=row, column=2, value=round(r['net_profit'], 2))
        row += 2

    ws2 = wb.create_sheet("Summary by Product")
    headers = ["Product", "Total Orders", "Positive Payouts", "Negative Payouts", "Net Payout", "Gross Profit", "Net Profit"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)
    row = 2
    for cat, agg in summary.items():
        ws2.cell(row=row, column=1, value=cat)
        ws2.cell(row=row, column=2, value=agg['total_orders'])
        ws2.cell(row=row, column=3, value=round(agg['total_positive'], 2))
        ws2.cell(row=row, column=4, value=round(agg['total_negative'], 2))
        ws2.cell(row=row, column=5, value=round(agg['net_payout'], 2))
        ws2.cell(row=row, column=6, value=round(agg['gross_profit'], 2))
        ws2.cell(row=row, column=7, value=round(agg['net_profit'], 2))
        row += 1
    row += 1
    ws2.cell(row=row, column=1, value="Grand Total Net Profit (before 5%)")
    ws2.cell(row=row, column=2, value=round(grand_total, 2))
    row += 1
    ws2.cell(row=row, column=1, value="Less 5% loss")
    ws2.cell(row=row, column=2, value=round(loss, 2))
    row += 1
    ws2.cell(row=row, column=1, value="FINAL NET PROFIT")
    ws2.cell(row=row, column=2, value=round(final_profit, 2))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_excel_detailed(detailed_data, total_net_profit_all):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Full Order Payouts")
    row = 1
    for item in detailed_data:
        ws.cell(row=row, column=1, value="Category: " + item['category'])
        row += 1
        ws.cell(row=row, column=1, value="Order Payouts")
        row += 1
        for payout in item['payouts']:
            ws.cell(row=row, column=1, value=round(payout, 2))
            row += 1
        ws.cell(row=row, column=1, value="Total Orders: " + str(item['count']))
        row += 1
        ws.cell(row=row, column=1, value="Total Payout: " + str(round(item['sum_payouts'], 2)))
        row += 1
        ws.cell(row=row, column=1, value="Cost per unit: " + str(round(item['cost'], 2)))
        row += 1
        ws.cell(row=row, column=1, value="Net Profit: " + str(round(item['net_profit'], 2)))
        row += 2
    loss = total_net_profit_all * 0.05
    final_profit = total_net_profit_all - loss
    ws.cell(row=row, column=1, value="Grand Total Net Profit (before 5%)")
    ws.cell(row=row, column=2, value=round(total_net_profit_all, 2))
    row += 1
    ws.cell(row=row, column=1, value="Less 5% loss")
    ws.cell(row=row, column=2, value=round(loss, 2))
    row += 1
    ws.cell(row=row, column=1, value="FINAL NET PROFIT")
    ws.cell(row=row, column=2, value=round(final_profit, 2))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf_summary(results, summary, grand_total, loss, final_profit):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    # Custom styles
    title_style = ParagraphStyle('TitleStyle', parent=styles['Title'], alignment=TA_CENTER, fontSize=18)
    heading_style = ParagraphStyle('HeadingStyle', parent=styles['Heading2'], alignment=TA_LEFT, fontSize=14)
    normal_style = styles['Normal']
    story = []

    # Title
    story.append(Paragraph("Profit Calculator – Summary Report", title_style))
    story.append(Spacer(1, 12))

    # Loop through each file
    for r in results:
        story.append(Paragraph(f"File: {r['filename']}", heading_style))
        story.append(Paragraph(f"Period: {r['period']}   |   Store: {r['store']}", normal_style))
        story.append(Spacer(1, 6))

        # Positive payouts table
        data = [["Positive Payouts", "Orders", "Total"]]
        for cat, d in r['positive'].items():
            data.append([cat, str(d['count']), f"{d['total']:.2f}"])
        t = Table(data, colWidths=[180, 80, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (2,0), (2,-1), 'RIGHT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('FONTNAME', (1,1), (-1,-1), 'Helvetica'),
        ]))
        story.append(t)
        story.append(Spacer(1, 6))

        # Negative payouts table
        data2 = [["Negative Payouts", "Orders", "Total"]]
        for cat, d in r['negative'].items():
            data2.append([cat, str(d['count']), f"{d['total']:.2f}"])
        t2 = Table(data2, colWidths=[180, 80, 100])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (2,0), (2,-1), 'RIGHT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
        ]))
        story.append(t2)
        story.append(Spacer(1, 6))

        # Gross profit and totals
        data3 = [["Gross Profit (after cost)", "Amount"]]
        for cat, profit in r['profits'].items():
            data3.append([cat, f"{profit:.2f}"])
        data3.append(["Total Gross Profit", f"{r['total_gross']:.2f}"])
        data3.append(["Total Negative Payouts", f"{r['total_neg']:.2f}"])
        data3.append(["Total Payout (before cost)", f"{r['total_payout']:.2f}"])
        data3.append(["Net Profit (after cost)", f"{r['net_profit']:.2f}"])
        t3 = Table(data3, colWidths=[250, 100])
        t3.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
        ]))
        story.append(t3)
        story.append(Spacer(1, 12))

    # Product summary
    story.append(Paragraph("Summary by Product (all files)", heading_style))
    story.append(Spacer(1, 6))
    data4 = [["Product", "Orders", "Positive", "Negative", "Net Payout", "Gross Profit", "Net Profit"]]
    for cat, agg in summary.items():
        data4.append([
            cat,
            str(agg['total_orders']),
            f"{agg['total_positive']:.2f}",
            f"{agg['total_negative']:.2f}",
            f"{agg['net_payout']:.2f}",
            f"{agg['gross_profit']:.2f}",
            f"{agg['net_profit']:.2f}"
        ])
    t4 = Table(data4, colWidths=[100, 50, 70, 70, 70, 70, 70])
    t4.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    story.append(t4)
    story.append(Spacer(1, 12))

    # Final totals
    story.append(Paragraph(f"Grand Total Net Profit (before 5%): {grand_total:.2f}", heading_style))
    story.append(Paragraph(f"Less 5% loss: -{loss:.2f}", normal_style))
    story.append(Paragraph(f"FINAL NET PROFIT: {final_profit:.2f}", title_style))

    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_pdf_detailed(detailed_data, total_net_profit_all):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Title'], alignment=TA_CENTER, fontSize=18)
    heading_style = ParagraphStyle('HeadingStyle', parent=styles['Heading2'], alignment=TA_LEFT, fontSize=14)
    normal_style = styles['Normal']
    story = []

    story.append(Paragraph("Profit Calculator – Detailed Order Payouts", title_style))
    story.append(Spacer(1, 12))

    for item in detailed_data:
        story.append(Paragraph(f"Category: {item['category']}", heading_style))
        story.append(Spacer(1, 4))
        # Payouts list as a table
        payouts_data = [["Order Payouts"]]
        for p in item['payouts']:
            payouts_data.append([f"{p:.2f}"])
        t = Table(payouts_data, colWidths=[100])
        t.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('FONTNAME', (1,0), (-1,-1), 'Helvetica'),
        ]))
        story.append(t)
        story.append(Spacer(1, 6))
        summary_data = [
            ["Total Orders", str(item['count'])],
            ["Total Payout", f"{item['sum_payouts']:.2f}"],
            ["Cost per unit", f"{item['cost']:.2f}"],
            ["Net Profit", f"{item['net_profit']:.2f}"]
        ]
        t2 = Table(summary_data, colWidths=[150, 100])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
        ]))
        story.append(t2)
        story.append(Spacer(1, 12))

    # Final totals with 5% deduction
    loss = total_net_profit_all * 0.05
    final_profit = total_net_profit_all - loss
    story.append(Paragraph(f"Grand Total Net Profit (before 5%): {total_net_profit_all:.2f}", heading_style))
    story.append(Paragraph(f"Less 5% loss: -{loss:.2f}", normal_style))
    story.append(Paragraph(f"FINAL NET PROFIT: {final_profit:.2f}", title_style))

    doc.build(story)
    buffer.seek(0)
    return buffer

# ---------- Flask Routes ----------
@app.route('/', methods=['GET', 'POST'])
def index():
    products = load_products()
    if request.method == 'POST':
        files = request.files.getlist('files')
        action = request.form.get('action')

        results = []
        grand_total = 0.0
        filenames = []
        summary = defaultdict(lambda: {'total_orders': 0, 'total_positive': 0.0, 'total_negative': 0.0,
                                       'net_payout': 0.0, 'gross_profit': 0.0, 'net_profit': 0.0})
        detailed_orders = defaultdict(list)
        detailed_costs = {}

        for f in files:
            if f.filename.endswith('.csv'):
                filepath = os.path.join('uploads', f.filename)
                os.makedirs('uploads', exist_ok=True)
                f.save(filepath)
                result, error = process_file(filepath)
                if error:
                    continue
                filenames.append(f.filename)

                pos_formatted = {}
                neg_formatted = {}
                for cat, vals in result['positive_payouts'].items():
                    pos_formatted[cat] = {'count': len(vals), 'total': sum(vals)}
                    summary[cat]['total_positive'] += sum(vals)
                    summary[cat]['total_orders'] += len(vals)
                for cat, vals in result['negative_payouts'].items():
                    neg_formatted[cat] = {'count': len(vals), 'total': sum(vals)}
                    summary[cat]['total_negative'] += sum(vals)
                for cat, profit in result['category_positive_profit'].items():
                    summary[cat]['gross_profit'] += profit

                for cat, vals in result['positive_payouts'].items():
                    detailed_orders[cat].extend(vals)
                for cat, vals in result['negative_payouts'].items():
                    detailed_orders[cat].extend(vals)
                for cat in result['category_positive_profit'].keys():
                    detailed_costs[cat] = get_cost(cat, products)

                results.append({
                    'filename': f.filename,
                    'period': result['period'],
                    'store': result['store'],
                    'positive': pos_formatted,
                    'negative': neg_formatted,
                    'profits': result['category_positive_profit'],
                    'total_gross': result['total_positive_profit'],
                    'total_neg': result['total_negative'],
                    'total_payout': result['total_payout'],
                    'net_profit': result['net_profit']
                })
                grand_total += result['net_profit']

        for cat in summary:
            summary[cat]['net_payout'] = summary[cat]['total_positive'] + summary[cat]['total_negative']
            summary[cat]['net_profit'] = summary[cat]['gross_profit'] + summary[cat]['total_negative']

        loss = grand_total * 0.05
        final_profit = grand_total - loss

        session['summary_data'] = {
            'results': results,
            'summary': dict(summary),
            'grand_total': grand_total,
            'loss': loss,
            'final_profit': final_profit
        }

        if action == 'calculate_full':
            detailed_data = []
            total_net_profit_all = 0.0
            for cat, payouts in detailed_orders.items():
                count = len(payouts)
                sum_payouts = sum(payouts)
                cost = detailed_costs.get(cat, 0)
                net_profit = sum_payouts - (count * cost)
                total_net_profit_all += net_profit
                detailed_data.append({
                    'category': cat,
                    'count': count,
                    'payouts': payouts,
                    'sum_payouts': sum_payouts,
                    'cost': cost,
                    'net_profit': net_profit
                })
            detailed_data.sort(key=lambda x: x['category'])
            session['detailed_data'] = {
                'detailed_data': detailed_data,
                'total_net_profit_all': total_net_profit_all
            }
            session.modified = True
            return render_template_string(
                HTML_TEMPLATE,
                results=None,
                files=filenames,
                products=products,
                summary=None,
                grand_total=None,
                loss=None,
                final_profit=None,
                has_results=False,
                detailed_data=detailed_data,
                total_net_profit_all=total_net_profit_all,
                action='calculate_full',
                has_detailed=True
            )
        else:
            session.modified = True
            return render_template_string(
                HTML_TEMPLATE,
                results=results,
                files=filenames,
                products=products,
                summary=summary,
                grand_total=grand_total,
                loss=loss,
                final_profit=final_profit,
                has_results=True,
                detailed_data=None,
                total_net_profit_all=None,
                action='calculate',
                has_detailed=False
            )

    return render_template_string(HTML_TEMPLATE, results=None, files=None, products=products, summary=None,
                                   grand_total=None, loss=None, final_profit=None, has_results=False,
                                   detailed_data=None, total_net_profit_all=None, action=None, has_detailed=False)

# ---------- Export routes ----------
@app.route('/export/excel')
def export_excel():
    if session.get('detailed_data'):
        data = session['detailed_data']
        output = generate_excel_detailed(data['detailed_data'], data['total_net_profit_all'])
        return send_file(output, as_attachment=True, download_name='full_order_payouts.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif session.get('summary_data'):
        data = session['summary_data']
        output = generate_excel_summary(data['results'], data['summary'], data['grand_total'], data['loss'], data['final_profit'])
        return send_file(output, as_attachment=True, download_name='profit_summary.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        return "No results to export. Please calculate first.", 400

@app.route('/export/pdf')
def export_pdf():
    if session.get('detailed_data'):
        data = session['detailed_data']
        output = generate_pdf_detailed(data['detailed_data'], data['total_net_profit_all'])
        return send_file(output, as_attachment=True, download_name='full_order_payouts.pdf', mimetype='application/pdf')
    elif session.get('summary_data'):
        data = session['summary_data']
        output = generate_pdf_summary(data['results'], data['summary'], data['grand_total'], data['loss'], data['final_profit'])
        return send_file(output, as_attachment=True, download_name='profit_summary.pdf', mimetype='application/pdf')
    else:
        return "No results to export. Please calculate first.", 400

# ---------- Product management routes ----------
@app.route('/update_cost', methods=['POST'])
def update_cost():
    data = request.json
    category = data.get('category')
    cost = data.get('cost')
    if not category or cost is None:
        return jsonify({'success': False, 'error': 'Missing data'})
    products = load_products()
    if category not in products:
        return jsonify({'success': False, 'error': 'Category not found'})
    products[category]['cost'] = cost
    save_products(products)
    return jsonify({'success': True})

@app.route('/delete_product', methods=['POST'])
def delete_product():
    data = request.json
    category = data.get('category')
    if not category:
        return jsonify({'success': False, 'error': 'Missing category'})
    products = load_products()
    if category not in products:
        return jsonify({'success': False, 'error': 'Category not found'})
    session['deleted_product'] = {
        'category': category,
        'cost': products[category]['cost'],
        'keywords': products[category]['keywords']
    }
    del products[category]
    save_products(products)
    return jsonify({'success': True})

@app.route('/undo_delete')
def undo_delete():
    deleted = session.pop('deleted_product', None)
    if deleted:
        products = load_products()
        products[deleted['category']] = {'cost': deleted['cost'], 'keywords': deleted['keywords']}
        save_products(products)
    return redirect('/')

@app.route('/clear_undo', methods=['POST'])
def clear_undo():
    session.pop('deleted_product', None)
    return jsonify({'success': True})

@app.route('/add_product', methods=['POST'])
def add_product():
    data = request.json
    name = data.get('name')
    cost = data.get('cost')
    keyword = data.get('keyword', '')
    if not name or cost is None:
        return jsonify({'success': False, 'error': 'Missing name or cost'})
    products = load_products()
    if name in products:
        return jsonify({'success': False, 'error': 'Product already exists'})
    if not keyword:
        keyword = name
    products[name] = {'cost': cost, 'keywords': [keyword]}
    save_products(products)
    return jsonify({'success': True})

# ---------- HTML Template (unchanged except undo alert fix) ----------
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Profit Calculator</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { background: #f0f2f5; padding: 20px; }
        .card { border-radius: 16px; box-shadow: 0 4px 12px rgba(0,0,0,0.08); margin-bottom: 20px; }
        .card-header { background: #fff; border-bottom: 1px solid #eee; font-weight: bold; }
        .product-item { display: flex; justify-content: space-between; align-items: center; padding: 8px 0; border-bottom: 1px solid #f0f0f0; }
        .product-item:last-child { border-bottom: none; }
        .cost-input { width: 80px; display: inline-block; }
        .result-box { background: #f8f9fa; border-radius: 12px; padding: 15px; margin-bottom: 15px; }
        .result-box h5 { margin-top: 0; }
        .positive { color: #28a745; }
        .negative { color: #dc3545; }
        .total-profit { font-size: 1.2rem; font-weight: bold; }
        .final-profit { background: #2c3e50; color: white; padding: 20px; border-radius: 16px; text-align: center; margin-top: 20px; }
        .final-profit .amount { font-size: 2.5rem; font-weight: bold; }
        .btn-sm { font-size: 0.8rem; }
        .keyword-badge { background: #e9ecef; padding: 2px 8px; border-radius: 12px; font-size: 0.7rem; margin-left: 5px; }
        .summary-box { background: #e8f4fd; border-radius: 12px; padding: 15px; margin-bottom: 10px; border-left: 4px solid #0d6efd; }
        .summary-box h6 { margin-top: 0; color: #0d6efd; }
        .undo-btn { background: #ffc107; color: #000; border: none; padding: 4px 12px; border-radius: 20px; font-size: 0.8rem; margin-left: 10px; }
        .export-buttons { display: flex; gap: 10px; justify-content: center; margin-top: 15px; flex-wrap: wrap; }
        .final-payout-box { background: #28a745; color: white; padding: 8px 16px; border-radius: 8px; display: inline-block; font-weight: bold; font-size: 1.4rem; }
        .total-payout-box { background: #17a2b8; color: white; padding: 8px 16px; border-radius: 8px; display: inline-block; font-weight: bold; font-size: 1.2rem; }
        .store-badge { background: #6c757d; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.9rem; }
        .designer-credit { text-align: center; color: #6c757d; margin-top: -10px; margin-bottom: 20px; font-size: 0.9rem; }
        .payout-list { max-height: 300px; overflow-y: auto; background: #f8f9fa; padding: 10px; border-radius: 8px; }
        .payout-item { font-family: monospace; }
        .detail-summary { background: #e9ecef; padding: 10px; border-radius: 8px; margin-top: 10px; }
    </style>
</head>
<body>
<div class="container">
    <h1 class="text-center mb-0">📊 Profit Calculator</h1>
    <p class="designer-credit">Designed by Shahar Yar</p>

    <!-- Undo deletion alert -->
    {% if session.deleted_product %}
    <div id="undo-alert" class="alert alert-warning alert-dismissible fade show" role="alert">
        <strong>Product "{{ session.deleted_product.category }}" deleted.</strong>
        <a href="{{ url_for('undo_delete') }}" class="btn btn-sm btn-outline-dark ms-2">↩ Undo</a>
        <button type="button" class="btn-close" data-bs-dismiss="alert" aria-label="Close"></button>
    </div>
    {% endif %}

    <!-- Product Management -->
    <div class="card">
        <div class="card-header">📦 Product Costs</div>
        <div class="card-body">
            <div id="product-list">
                {% for cat, data in products.items() %}
                <div class="product-item" data-category="{{ cat }}">
                    <div>
                        <strong>{{ cat }}</strong>
                        <span class="keyword-badge">{{ data.keywords|join(', ') }}</span>
                    </div>
                    <div>
                        <span id="cost-display-{{ loop.index }}">{{ "%.2f"|format(data.cost) }}</span>
                        <input type="number" step="0.01" class="form-control form-control-sm cost-input d-none" id="cost-input-{{ loop.index }}" value="{{ data.cost }}">
                        <button class="btn btn-sm btn-outline-primary edit-btn" data-index="{{ loop.index }}">Edit</button>
                        <button class="btn btn-sm btn-outline-danger delete-btn" data-category="{{ cat }}">Delete</button>
                        <button class="btn btn-sm btn-success save-btn d-none" data-index="{{ loop.index }}">Save</button>
                    </div>
                </div>
                {% endfor %}
            </div>
            <hr>
            <h6>Add New Product</h6>
            <div class="row g-2">
                <div class="col-md-4">
                    <input type="text" id="new-name" class="form-control" placeholder="Product Name">
                </div>
                <div class="col-md-3">
                    <input type="number" step="0.01" id="new-cost" class="form-control" placeholder="Cost">
                </div>
                <div class="col-md-3">
                    <input type="text" id="new-keyword" class="form-control" placeholder="Keyword (optional)">
                </div>
                <div class="col-md-2">
                    <button class="btn btn-primary w-100" id="add-product-btn">Add</button>
                </div>
            </div>
            <div id="add-message" class="mt-2"></div>
        </div>
    </div>

    <!-- File Upload -->
    <div class="card">
        <div class="card-header">📂 Upload CSV Files</div>
        <div class="card-body">
            <form method="POST" enctype="multipart/form-data" id="upload-form">
                <input type="file" name="files" multiple accept=".csv" class="form-control mb-2" required>
                <div class="row g-2">
                    <div class="col-md-6">
                        <button type="submit" name="action" value="calculate" class="btn btn-success w-100">Calculate</button>
                    </div>
                    <div class="col-md-6">
                        <button type="submit" name="action" value="calculate_full" class="btn btn-primary w-100">Calculate Full</button>
                    </div>
                </div>
            </form>
            {% if files %}
            <div class="mt-2"><strong>Files processed:</strong> {{ files|join(', ') }}</div>
            {% endif %}
        </div>
    </div>

    <!-- ============================================================ -->
    <!-- 1) SUMMARY VIEW (when 'Calculate' is clicked)                  -->
    <!-- ============================================================ -->
    {% if results and action == 'calculate' %}
    <div class="card">
        <div class="card-header">📈 Results per File (Summary)</div>
        <div class="card-body">
            {% for r in results %}
            <div class="result-box">
                <h5>
                    📁 {{ r.filename }}
                    <span class="store-badge">🏪 {{ r.store }}</span>
                    <small class="text-muted">({{ r.period }})</small>
                </h5>
                <div class="row">
                    <div class="col-md-6">
                        <strong>Positive payouts:</strong>
                        {% for cat, data in r.positive.items() %}
                        <div><span class="positive">{{ cat }}</span>: {{ data.count }} orders, total {{ "%.2f"|format(data.total) }}</div>
                        {% endfor %}
                    </div>
                    <div class="col-md-6">
                        <strong>Negative payouts:</strong>
                        {% for cat, data in r.negative.items() %}
                        <div><span class="negative">{{ cat }}</span>: {{ data.count }} orders, total {{ "%.2f"|format(data.total) }}</div>
                        {% endfor %}
                    </div>
                </div>
                <div class="row mt-2">
                    <div class="col-md-6">
                        <strong>Gross profit (after cost):</strong>
                        {% for cat, profit in r.profits.items() %}
                        <div>{{ cat }}: {{ "%.2f"|format(profit) }}</div>
                        {% endfor %}
                    </div>
                    <div class="col-md-6">
                        <div>Total gross profit: <span class="positive">{{ "%.2f"|format(r.total_gross) }}</span></div>
                        <div>Total negative payouts: <span class="negative">{{ "%.2f"|format(r.total_neg) }}</span></div>
                    </div>
                </div>
                <div class="mt-3">
                    <strong>Total Orders:</strong>
                    {% set total_orders = 0 %}
                    {% for cat, data in r.positive.items() %}{% set total_orders = total_orders + data.count %}{% endfor %}
                    {% for cat, data in r.negative.items() %}{% set total_orders = total_orders + data.count %}{% endfor %}
                    {{ total_orders }}
                    &nbsp;&nbsp;|&nbsp;&nbsp;
                    <span class="total-payout-box">📊 Total Payout (before cost): {{ "%.2f"|format(r.total_payout) }}</span>
                    &nbsp;&nbsp;|&nbsp;&nbsp;
                    <span class="final-payout-box">💰 Net Profit (after cost): {{ "%.2f"|format(r.net_profit) }}</span>
                </div>
            </div>
            {% endfor %}

            <!-- Summary by Product (aggregated) -->
            <div class="card mt-4">
                <div class="card-header">📊 Summary by Product (all files)</div>
                <div class="card-body">
                    {% for cat, agg in summary.items() %}
                    <div class="summary-box">
                        <h6>{{ cat }}</h6>
                        <div class="row">
                            <div class="col-sm-3">Total Orders: <strong>{{ agg.total_orders }}</strong></div>
                            <div class="col-sm-3">Positive Payouts: <span class="positive">{{ "%.2f"|format(agg.total_positive) }}</span></div>
                            <div class="col-sm-3">Negative Payouts: <span class="negative">{{ "%.2f"|format(agg.total_negative) }}</span></div>
                            <div class="col-sm-3">Net Payout: <span class="total-profit">{{ "%.2f"|format(agg.net_payout) }}</span></div>
                        </div>
                        <div>Gross profit (after cost): {{ "%.2f"|format(agg.gross_profit) }}</div>
                        <div><strong>Net Profit:</strong> <span class="total-profit">{{ "%.2f"|format(agg.net_profit) }}</span></div>
                    </div>
                    {% endfor %}
                </div>
            </div>

            <!-- Final combined -->
            <div class="final-profit">
                <div>Grand Total Net Profit (before 5%)</div>
                <div class="amount">{{ "%.2f"|format(grand_total) }}</div>
                <div>Less 5% loss: -{{ "%.2f"|format(loss) }}</div>
                <div style="font-size: 1.8rem; margin-top: 10px;">💰 FINAL NET PROFIT: {{ "%.2f"|format(final_profit) }}</div>
            </div>

            <!-- Export buttons -->
            <div class="export-buttons">
                <a href="{{ url_for('export_excel') }}" class="btn btn-success btn-lg">📥 Download Excel</a>
                <a href="{{ url_for('export_pdf') }}" class="btn btn-danger btn-lg">📥 Download PDF</a>
            </div>
        </div>
    </div>
    {% endif %}

    <!-- ============================================================ -->
    <!-- 2) DETAILED VIEW (when 'Calculate Full' is clicked)           -->
    <!-- ============================================================ -->
    {% if detailed_data and action == 'calculate_full' %}
    <div class="card">
        <div class="card-header">📋 Full Order Payouts per Product</div>
        <div class="card-body">
            {% for item in detailed_data %}
            <div class="result-box">
                <h5>{{ item.category }}</h5>
                <div class="payout-list">
                    {% for payout in item.payouts %}
                    <div class="payout-item">{{ "%.2f"|format(payout) }}</div>
                    {% endfor %}
                </div>
                <div class="detail-summary">
                    <div>Total Orders: <strong>{{ item.count }}</strong></div>
                    <div>Total Payout: <span class="total-profit">{{ "%.2f"|format(item.sum_payouts) }}</span></div>
                    <div>Cost per unit: {{ "%.2f"|format(item.cost) }}</div>
                    <div><strong>Net Profit:</strong> <span class="final-payout-box">{{ "%.2f"|format(item.net_profit) }}</span></div>
                </div>
            </div>
            {% endfor %}
            <div class="final-profit" style="margin-top: 20px;">
                <div>Grand Total Net Profit (before 5%)</div>
                <div class="amount">{{ "%.2f"|format(total_net_profit_all) }}</div>
                {% set loss_detailed = total_net_profit_all * 0.05 %}
                {% set final_profit_detailed = total_net_profit_all - loss_detailed %}
                <div>Less 5% loss: -{{ "%.2f"|format(loss_detailed) }}</div>
                <div style="font-size: 1.8rem; margin-top: 10px;">💰 FINAL NET PROFIT: {{ "%.2f"|format(final_profit_detailed) }}</div>
            </div>

            <!-- Export buttons for detailed view -->
            <div class="export-buttons">
                <a href="{{ url_for('export_excel') }}" class="btn btn-success btn-lg">📥 Download Excel</a>
                <a href="{{ url_for('export_pdf') }}" class="btn btn-danger btn-lg">📥 Download PDF</a>
            </div>
        </div>
    </div>
    {% endif %}

</div>

<script>
    // Clear undo session when alert is closed
    document.addEventListener('DOMContentLoaded', function() {
        var undoAlert = document.getElementById('undo-alert');
        if (undoAlert) {
            undoAlert.addEventListener('closed.bs.alert', function () {
                fetch('/clear_undo', { method: 'POST' });
            });
        }
    });

    // Product management JavaScript (unchanged)
    document.querySelectorAll('.edit-btn').forEach(btn => {
        btn.addEventListener('click', function() {
            const idx = this.dataset.index;
            const display = document.getElementById('cost-display-' + idx);
            const input = document.getElementById('cost-input-' + idx);
            const saveBtn = document.querySelector('.save-btn[data-index="' + idx + '"]');
            display.classList.add('d-none');
            input.classList.remove('d-none');
            this.classList.add('d-none');
            saveBtn.classList.remove('d-none');
        });
    });

    document.querySelectorAll('.save-btn').forEach(btn => {
        btn.addEventListener('click', function() {
            const idx = this.dataset.index;
            const display = document.getElementById('cost-display-' + idx);
            const input = document.getElementById('cost-input-' + idx);
            const editBtn = document.querySelector('.edit-btn[data-index="' + idx + '"]');
            const category = this.closest('.product-item').dataset.category;
            const newCost = parseFloat(input.value);
            if (isNaN(newCost) || newCost < 0) {
                alert('Please enter a valid cost.');
                return;
            }
            fetch('/update_cost', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({category: category, cost: newCost})
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    display.textContent = newCost.toFixed(2);
                    display.classList.remove('d-none');
                    input.classList.add('d-none');
                    editBtn.classList.remove('d-none');
                    this.classList.add('d-none');
                    location.reload();
                } else {
                    alert('Error updating cost.');
                }
            });
        });
    });

    document.querySelectorAll('.delete-btn').forEach(btn => {
        btn.addEventListener('click', function() {
            const category = this.dataset.category;
            if (!confirm('Are you sure you want to delete "' + category + '"?')) return;
            fetch('/delete_product', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({category: category})
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    location.reload();
                } else {
                    alert('Error deleting product.');
                }
            });
        });
    });

    document.getElementById('add-product-btn').addEventListener('click', function() {
        const name = document.getElementById('new-name').value.trim();
        const cost = parseFloat(document.getElementById('new-cost').value);
        const keyword = document.getElementById('new-keyword').value.trim();
        if (!name || isNaN(cost) || cost < 0) {
            document.getElementById('add-message').innerHTML = '<div class="alert alert-danger">Please enter a valid name and cost.</div>';
            return;
        }
        fetch('/add_product', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({name: name, cost: cost, keyword: keyword})
        })
        .then(response => response.json())
        .then(data => {
            if (data.success) {
                location.reload();
            } else {
                document.getElementById('add-message').innerHTML = '<div class="alert alert-danger">' + data.error + '</div>';
            }
        });
    });
</script>
</body>
</html>
'''

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)